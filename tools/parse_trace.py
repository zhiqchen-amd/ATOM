#!/usr/bin/env python3
"""Small, step-by-step ATOM trace parser.

Step 1: find the decode warmup window in the capture trace that corresponds to
the first decode event in the run trace.
"""

from __future__ import annotations

import argparse
import csv
import gzip
import json
import os
import re
from glob import glob
from typing import Any

SPECIAL_KERNEL_LAUNCH_NAMES = {"hipmemcpyasync"}


def load_events(path: str) -> list[dict[str, Any]]:
    opener = gzip.open if path.endswith(".gz") else open
    with opener(path, "rt", encoding="utf-8", errors="replace") as f:
        return json.load(f).get("traceEvents", [])


def event_end(event: dict[str, Any]) -> float:
    return float(event.get("ts", 0.0)) + float(event.get("dur", 0.0))


def is_kernel_launch(name: str) -> bool:
    normalized = name.lower()
    return (
        "launch" in normalized and "kernel" in normalized
    ) or normalized in SPECIAL_KERNEL_LAUNCH_NAMES


def short(text: Any, limit: int = 80) -> str:
    value = str(text)
    return value if len(value) <= limit else value[: limit - 3] + "..."


def model_name_from_trace(path: str) -> str | None:
    base = os.path.basename(path)
    if "_ts_" not in base:
        return None
    prefix = base.split("_ts_", 1)[0]
    if prefix.startswith("capture_graph_"):
        prefix = prefix[len("capture_graph_") :]
    return prefix or None


def find_capture_trace(run_trace: str) -> str | None:
    model_name = model_name_from_trace(run_trace)
    if not model_name:
        return None
    trace_dir = os.path.dirname(run_trace) or "."
    pattern = os.path.join(trace_dir, f"capture_graph_{model_name}_ts_*.pt.trace.json*")
    candidates = sorted(glob(pattern), key=os.path.getmtime, reverse=True)
    run_abs = os.path.abspath(run_trace)
    for candidate in candidates:
        if os.path.abspath(candidate) != run_abs:
            return candidate
    return None


def find_first_decode(events: list[dict[str, Any]]) -> dict[str, Any]:
    decodes = sorted(
        [
            event
            for event in events
            if event.get("ph") == "X"
            and event.get("cat") == "gpu_user_annotation"
            and str(event.get("name", "")).startswith("decode[")
        ],
        key=lambda event: event["ts"],
    )
    if not decodes:
        raise RuntimeError("No decode gpu_user_annotation found in run trace.")
    return decodes[0]


def decode_batch_size(decode_event: dict[str, Any]) -> int:
    match = re.search(r"bs=(\d+)", str(decode_event.get("name", "")))
    if not match:
        raise RuntimeError(
            f"Could not parse batch size from {decode_event.get('name')!r}"
        )
    return int(match.group(1))


def find_cpu_capture_graphs(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(
        [
            event
            for event in events
            if event.get("ph") == "X"
            and event.get("cat") == "user_annotation"
            and str(event.get("name", "")).startswith("capture_graph_bs_")
        ],
        key=lambda event: event["ts"],
    )


def find_capture_graph_for_bs(
    capture_events: list[dict[str, Any]], batch_size: int
) -> dict[str, Any]:
    graphs = find_cpu_capture_graphs(capture_events)
    if not graphs:
        raise RuntimeError(
            "No CPU capture_graph_bs_* annotations found in capture trace."
        )
    target_name = f"capture_graph_bs_{batch_size}"
    for graph in graphs:
        if graph.get("name") == target_name:
            return graph
    raise RuntimeError(f"No {target_name} found in capture trace.")


def warmup_window_for_graph(
    capture_events: list[dict[str, Any]], target_graph: dict[str, Any]
) -> tuple[float, float]:
    """Return [previous_capture_graph_end, target_capture_graph_start)."""
    start = 0.0
    for graph in find_cpu_capture_graphs(capture_events):
        if graph is target_graph:
            return start, float(target_graph["ts"])
        start = max(start, event_end(graph))
    raise RuntimeError("Target capture graph was not in capture graph list.")


def count_events_in_window(
    events: list[dict[str, Any]], start: float, end: float
) -> dict[str, int]:
    counts = {"duration": 0, "user_annotation": 0, "cuda_runtime": 0, "kernel": 0}
    for event in events:
        if event.get("ph") != "X":
            continue
        ts = float(event.get("ts", 0.0))
        if not (start <= ts < end):
            continue
        counts["duration"] += 1
        cat = event.get("cat")
        if cat in counts:
            counts[cat] += 1
    return counts


def build_correlation_index(
    events: list[dict[str, Any]], start: float, end: float
) -> tuple[dict[Any, dict[str, Any]], dict[Any, dict[str, Any]]]:
    launches: dict[Any, dict[str, Any]] = {}
    kernels: dict[Any, dict[str, Any]] = {}
    for event in events:
        if event.get("ph") != "X":
            continue
        ts = float(event.get("ts", 0.0))
        if not (start <= ts < end):
            continue
        corr = (event.get("args") or {}).get("correlation")
        if corr is None:
            continue
        if event.get("cat") == "cuda_runtime" and is_kernel_launch(
            str(event.get("name", ""))
        ):
            launches.setdefault(corr, event)
        elif event.get("cat") == "kernel":
            kernels.setdefault(corr, event)
    return launches, kernels


def containing_annotations(
    event: dict[str, Any], annotations: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    start = float(event["ts"])
    end = event_end(event)
    return [
        ann
        for ann in annotations
        if ann.get("pid") == event.get("pid")
        and ann.get("tid") == event.get("tid")
        and float(ann.get("ts", 0.0)) <= start
        and end <= event_end(ann)
    ]


def is_compiled_graph_tag(name: str) -> bool:
    return name.startswith("## Call CompiledFxGraph")


def cpu_fallback_tag_for_compiled_kernel(
    kernel: dict[str, Any],
    launch_by_corr: dict[Any, dict[str, Any]],
    cpu_events: list[dict[str, Any]],
) -> str | None:
    corr = (kernel.get("args") or {}).get("correlation")
    launch = launch_by_corr.get(corr)
    if launch is None:
        return None

    containers = containing_annotations(launch, cpu_events)
    # Include cpu_op parents as well as user annotations, then pick the largest
    # non-CompiledFxGraph container. This maps tiny compiled graph kernels such
    # as FillFunctor copies back to semantic CPU ops like aiter::all_reduce_.
    start = float(launch["ts"])
    end = event_end(launch)
    containers.extend(
        event
        for event in cpu_events
        if event.get("cat") == "cpu_op"
        and event.get("pid") == launch.get("pid")
        and event.get("tid") == launch.get("tid")
        and float(event.get("ts", 0.0)) <= start
        and end <= event_end(event)
    )
    candidates = [
        event
        for event in containers
        if not is_compiled_graph_tag(str(event.get("name", "")))
    ]
    if not candidates:
        return None
    return str(
        max(candidates, key=lambda event: float(event.get("dur", 0.0))).get("name")
    )


def gpu_tag_for_kernel(
    kernel: dict[str, Any],
    gpu_annotations: list[dict[str, Any]],
    launch_by_corr: dict[Any, dict[str, Any]],
    cpu_events: list[dict[str, Any]],
) -> str:
    containers = containing_annotations(kernel, gpu_annotations)
    if not containers:
        return "<no gpu tag>"
    tag = str(
        min(containers, key=lambda event: float(event.get("dur", 0.0))).get("name")
    )
    if is_compiled_graph_tag(tag):
        fallback = cpu_fallback_tag_for_compiled_kernel(
            kernel, launch_by_corr, cpu_events
        )
        if fallback:
            return fallback
    return tag


def build_warmup_mapping(
    capture_events: list[dict[str, Any]], start: float, end: float
) -> list[dict[str, Any]]:
    """Build the internal decode warmup mapping.

    Each row is intentionally minimal:
      - module: resolved CPU/GPU tag name
      - kernel: GPU kernel name
      - stream: GPU stream id

    This mapping is the attribution source for later replay-duration matching;
    it is not meant to be emitted as the final user-facing breakdown.
    """
    launch_by_corr, _ = build_correlation_index(capture_events, start, end)
    cpu_events = [
        event
        for event in capture_events
        if event.get("ph") == "X"
        and start <= float(event.get("ts", 0.0)) < end
        and event.get("cat") in {"user_annotation", "cpu_op"}
    ]
    gpu_annotations = [
        event
        for event in capture_events
        if event.get("ph") == "X"
        and start <= float(event.get("ts", 0.0)) < end
        and event.get("cat") == "gpu_user_annotation"
    ]
    kernels = sorted(
        [
            event
            for event in capture_events
            if event.get("ph") == "X"
            and start <= float(event.get("ts", 0.0)) < end
            and event.get("cat") == "kernel"
        ],
        key=lambda event: event["ts"],
    )
    mapping: list[dict[str, Any]] = []
    for kernel in kernels:
        args = kernel.get("args") or {}
        mapping.append(
            {
                "module": gpu_tag_for_kernel(
                    kernel, gpu_annotations, launch_by_corr, cpu_events
                ),
                "kernel": str(kernel.get("name", "")),
                "stream": args.get("stream"),
            }
        )
    return mapping


def print_first_warmup_mappings(mapping: list[dict[str, Any]], limit: int) -> None:
    print("")
    print(f"First {limit} warmup mappings:")
    print("| # | module/tag | stream | kernel |")
    print("|---:|---|---:|---|")
    for idx, item in enumerate(mapping[:limit]):
        print(
            f"| {idx} | `{short(item['module'], 55)}` | {item['stream']} | "
            f"`{short(item['kernel'], 85)}` |"
        )


def decode_gpu_window(
    run_events: list[dict[str, Any]], decode_event: dict[str, Any]
) -> tuple[float, float]:
    """Return the GPU annotation time range for the selected decode event.

    We intentionally use the GPU-side annotation range here: the final CSV is
    for observed replay GPU kernels. Kernels that fall just outside this range
    are not included in this first formal path.
    """
    external_id = (decode_event.get("args") or {}).get("External id")
    gpu_decodes = [
        event
        for event in run_events
        if event.get("ph") == "X"
        and event.get("cat") == "gpu_user_annotation"
        and (event.get("args") or {}).get("External id") == external_id
    ]
    if not gpu_decodes:
        # Fallback for traces without GPU annotation projection.
        return float(decode_event["ts"]), event_end(decode_event)
    return min(float(event["ts"]) for event in gpu_decodes), max(
        event_end(event) for event in gpu_decodes
    )


def replay_kernels_in_window(
    run_events: list[dict[str, Any]], start: float, end: float
) -> list[dict[str, Any]]:
    return sorted(
        [
            event
            for event in run_events
            if event.get("ph") == "X"
            and event.get("cat") == "kernel"
            and start <= float(event.get("ts", 0.0)) < end
        ],
        key=lambda event: event["ts"],
    )


def remap_streams(replay_kernels: list[dict[str, Any]]) -> dict[Any, int]:
    """Map real replay stream ids to compact 1..N ids by numeric order."""
    streams = sorted(
        {(event.get("args") or {}).get("stream") for event in replay_kernels},
        key=lambda value: (value is None, value),
    )
    return {stream: idx + 1 for idx, stream in enumerate(streams)}


LAYER_RE = re.compile(r"(^|\.)layers\.(\d+)\.")


def module_layer(module: str) -> int | None:
    match = LAYER_RE.search(module)
    return int(match.group(2)) if match else None


def normalize_layer_module(module: str) -> str:
    return re.sub(r"(^|\.)layers\.\d+\.", r"\1layers.*.", module)


def layer_group_label(layers: list[int]) -> str:
    layers = sorted(layers)
    if not layers:
        return "layers <empty>"
    if len(layers) == 1:
        return f"layer {layers[0]}"
    if layers == list(range(layers[0], layers[-1] + 1)):
        return f"layers {layers[0]}-{layers[-1]}"
    if len(layers) <= 8:
        return "layers " + ",".join(str(layer) for layer in layers)
    return f"layers {layers[0]},{layers[1]},...,{layers[-1]} ({len(layers)} layers)"


def warmup_item_owner_layers(warmup_mapping: list[dict[str, Any]]) -> list[int | None]:
    """Assign warmup rows to layer parse windows.

    A few runtime helper kernels, such as maybe_dual_stream_forward, appear as
    non-layer tags between two chunks of the same layer.  They should still be
    consumed while parsing that layer; their original module tag is preserved in
    the final row.
    """
    direct_layers = [module_layer(item["module"]) for item in warmup_mapping]
    owners: list[int | None] = []
    for idx, layer in enumerate(direct_layers):
        if layer is not None:
            owners.append(layer)
            continue

        prev_layer = None
        for prev_idx in range(idx - 1, -1, -1):
            if direct_layers[prev_idx] is not None:
                prev_layer = direct_layers[prev_idx]
                break

        next_layer = None
        for next_idx in range(idx + 1, len(direct_layers)):
            if direct_layers[next_idx] is not None:
                next_layer = direct_layers[next_idx]
                break

        owners.append(
            prev_layer if prev_layer is not None and prev_layer == next_layer else None
        )
    return owners


def warmup_parse_blocks(
    warmup_mapping: list[dict[str, Any]],
) -> list[tuple[int | None, list[tuple[int, dict[str, Any]]]]]:
    owners = warmup_item_owner_layers(warmup_mapping)
    blocks: list[tuple[int | None, list[tuple[int, dict[str, Any]]]]] = []
    for idx, (owner, item) in enumerate(zip(owners, warmup_mapping, strict=True)):
        if not blocks or blocks[-1][0] != owner:
            blocks.append((owner, []))
        blocks[-1][1].append((idx, item))
    return blocks


def primary_replay_stream(replay_kernels: list[dict[str, Any]]) -> Any:
    counts: dict[Any, int] = {}
    first_ts: dict[Any, float] = {}
    for event in replay_kernels:
        stream = (event.get("args") or {}).get("stream")
        counts[stream] = counts.get(stream, 0) + 1
        first_ts.setdefault(stream, float(event.get("ts", 0.0)))
    return max(counts, key=lambda stream: (counts[stream], -first_ts[stream]))


def consume_replay_stream_for_block(
    stream_events: list[dict[str, Any]],
    cursor: int,
    template: list[tuple[int, dict[str, Any]]],
    used_template_positions: set[int],
) -> tuple[int, list[tuple[int, dict[str, Any], dict[str, Any]]]]:
    """Consume one replay stream against one warmup layer block.

    The stream is matched event-by-event to the next compatible warmup kernel in
    this block, skipping template rows that ran on other replay streams.
    """
    rows: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
    template_pos = 0
    pos = cursor
    while pos < len(stream_events):
        replay = stream_events[pos]
        replay_kernel = str(replay.get("name", ""))
        matched_pos = None
        for idx in range(template_pos, len(template)):
            if idx in used_template_positions:
                continue
            if template[idx][1]["kernel"] == replay_kernel:
                matched_pos = idx
                break
        if matched_pos is None:
            break

        used_template_positions.add(matched_pos)
        template_pos = matched_pos + 1
        rows.append((template[matched_pos][0], template[matched_pos][1], replay))
        pos += 1
    return pos, rows


def match_replay_to_warmup(
    replay_kernels: list[dict[str, Any]], warmup_mapping: list[dict[str, Any]]
) -> list[dict[str, Any]]:
    """Match replay to warmup layer-by-layer without stream-id assumptions.

    Warmup mapping is treated as the semantic operator template.  For each layer
    block, the primary replay stream is consumed first, then the remaining
    streams consume the still-unmatched kernels in the same block.  This avoids
    assuming capture streams and CUDAGraph replay streams are one-to-one.
    """
    if not replay_kernels:
        return []

    replay_by_stream: dict[Any, list[dict[str, Any]]] = {}
    for event in replay_kernels:
        stream = (event.get("args") or {}).get("stream")
        replay_by_stream.setdefault(stream, []).append(event)

    main_stream = primary_replay_stream(replay_kernels)
    other_streams = [
        stream
        for stream in sorted(replay_by_stream, key=lambda value: (value is None, value))
        if stream != main_stream
    ]
    stream_order = [main_stream] + other_streams
    stream_cursors = {stream: 0 for stream in replay_by_stream}
    used_warmup_indices: set[int] = set()
    warmup_owner_layers = warmup_item_owner_layers(warmup_mapping)
    rows: list[dict[str, Any]] = []

    for _, template in warmup_parse_blocks(warmup_mapping):
        used_template_positions: set[int] = set()
        block_matches: list[tuple[int, dict[str, Any], dict[str, Any]]] = []
        for stream in stream_order:
            next_cursor, stream_matches = consume_replay_stream_for_block(
                replay_by_stream[stream],
                stream_cursors[stream],
                template,
                used_template_positions,
            )
            stream_cursors[stream] = next_cursor
            block_matches.extend(stream_matches)

        for warmup_idx, matched, replay in sorted(
            block_matches, key=lambda item: item[0]
        ):
            used_warmup_indices.add(warmup_idx)
            replay_stream = (replay.get("args") or {}).get("stream")
            rows.append(
                {
                    "warmup_index": warmup_idx,
                    "cpu_module": matched["module"],
                    "owner_layer": warmup_owner_layers[warmup_idx],
                    "kernel_name": str(replay.get("name", "")),
                    "stream": replay_stream,
                    "duration_us": float(replay.get("dur", 0.0)),
                }
            )

    leftovers: list[dict[str, Any]] = []
    for stream, events in replay_by_stream.items():
        leftovers.extend(events[stream_cursors[stream] :])

    unused_warmup_by_kernel: dict[str, list[tuple[int, dict[str, Any]]]] = {}
    for idx, item in enumerate(warmup_mapping):
        if idx in used_warmup_indices:
            continue
        unused_warmup_by_kernel.setdefault(item["kernel"], []).append((idx, item))

    for replay in sorted(leftovers, key=lambda event: float(event.get("ts", 0.0))):
        replay_stream = (replay.get("args") or {}).get("stream")
        replay_kernel = str(replay.get("name", ""))
        matched_items = unused_warmup_by_kernel.get(replay_kernel, [])
        matched_idx, matched = matched_items.pop(0) if matched_items else (-1, None)
        rows.append(
            {
                "warmup_index": matched_idx,
                "cpu_module": matched["module"] if matched else "<unmatched>",
                "owner_layer": (
                    warmup_owner_layers[matched_idx] if matched_idx >= 0 else None
                ),
                "kernel_name": replay_kernel,
                "stream": replay_stream,
                "duration_us": float(replay.get("dur", 0.0)),
            }
        )
    return sorted(
        rows,
        key=lambda row: (
            row["warmup_index"] < 0,
            row["warmup_index"],
        ),
    )


def build_grouped_breakdown_rows(
    rows: list[dict[str, Any]], stream_map: dict[Any, int], print_head: bool = False
) -> list[dict[str, Any]]:
    """Aggregate matched replay rows into layer-structure groups.

    Layer rows are grouped by identical per-layer operator sequence.  Each output
    row for a layer group is the average time for that operator position across
    layers in the group.  Non-layer and unmatched rows are aggregated by
    module/kernel/stream.
    """
    layer_rows: dict[int, list[dict[str, Any]]] = {}
    non_layer_accum: dict[tuple[str, str, int], tuple[float, int]] = {}
    unmatched_accum: dict[tuple[str, int], float] = {}

    for row in rows:
        stream_no = stream_map.get(row["stream"], 0)
        if row["cpu_module"] == "<unmatched>":
            key = (row["kernel_name"], stream_no)
            unmatched_accum[key] = unmatched_accum.get(key, 0.0) + row["duration_us"]
            continue

        layer = module_layer(row["cpu_module"])
        if layer is None:
            layer = row.get("owner_layer")
        normalized = normalize_layer_module(row["cpu_module"])
        normalized_row = {
            **row,
            "module_pattern": normalized,
            "stream_no": stream_no,
        }
        if layer is None:
            key = (normalized, row["kernel_name"], stream_no)
            total, order_key = non_layer_accum.get(key, (0.0, row["warmup_index"]))
            non_layer_accum[key] = (
                total + row["duration_us"],
                min(order_key, row["warmup_index"]),
            )
        else:
            layer_rows.setdefault(layer, []).append(normalized_row)

    grouped: list[dict[str, Any]] = []

    # Non-layer prologue/epilogue rows are hidden by default. They still
    # contribute to the full-decode denominator.
    if print_head:
        for (module, kernel, stream_no), (
            total_us,
            order_key,
        ) in non_layer_accum.items():
            grouped.append(
                {
                    "layer_group": "non_layer",
                    "module": module,
                    "kernel": kernel,
                    "stream_no": stream_no,
                    "time_us": total_us,
                    "order_key": order_key,
                    "group_order_key": order_key,
                }
            )

    # Layer groups by exact normalized operator sequence.  Stream id is kept as
    # output metadata, but does not decide whether two layers have the same
    # operator layout.
    signature_to_layers: dict[tuple[tuple[str, str], ...], list[int]] = {}
    for layer in sorted(layer_rows):
        items = layer_rows[layer]
        signature = tuple(
            (item["module_pattern"], item["kernel_name"]) for item in items
        )
        signature_to_layers.setdefault(signature, []).append(layer)

    for signature, layers in signature_to_layers.items():
        layer_count = len(layers)
        label = layer_group_label(layers)
        group_order_key = min(layer_rows[layer][0]["warmup_index"] for layer in layers)
        for idx, (module, kernel) in enumerate(signature):
            total = sum(layer_rows[layer][idx]["duration_us"] for layer in layers)
            stream_ids = [layer_rows[layer][idx]["stream_no"] for layer in layers]
            order_key = min(layer_rows[layer][idx]["warmup_index"] for layer in layers)
            stream_no: int | str
            if all(stream_id == stream_ids[0] for stream_id in stream_ids):
                stream_no = stream_ids[0]
            else:
                stream_no = ",".join(
                    str(stream_id) for stream_id in sorted(set(stream_ids))
                )
            grouped.append(
                {
                    "layer_group": label,
                    "module": module,
                    "kernel": kernel,
                    "stream_no": stream_no,
                    "time_us": total / layer_count,
                    "order_key": order_key,
                    "group_order_key": group_order_key,
                }
            )
        group_time_us = sum(
            sum(layer_rows[layer][idx]["duration_us"] for layer in layers) / layer_count
            for idx in range(len(signature))
        )
        grouped.append(
            {
                "layer_group": label,
                "module": "__group_total__",
                "kernel": "GROUP TOTAL",
                "stream_no": "",
                "time_us": group_time_us,
                "order_key": float("inf"),
                "group_order_key": group_order_key,
            }
        )

    # Keep unmatched in the same output file for follow-up instrumentation work.
    for (kernel, stream_no), total_us in unmatched_accum.items():
        grouped.append(
            {
                "layer_group": "unmatched",
                "module": "<unmatched>",
                "kernel": kernel,
                "stream_no": stream_no,
                "time_us": total_us,
                "order_key": float("inf"),
                "group_order_key": float("inf"),
            }
        )

    return sorted(grouped, key=lambda row: (row["group_order_key"], row["order_key"]))


DECODE_BREAKDOWN_HEADER = [
    "layer_group",
    "module/tag",
    "kernel",
    "stream_id",
    "time_us",
    "percent_of_full_decode_forward",
]
XLSX_KERNEL_DISPLAY_LIMIT = 120


def decode_breakdown_values(
    rows: list[dict[str, Any]], full_decode_us: float
) -> list[list[Any]]:
    values: list[list[Any]] = []
    for row in rows:
        percent = row["time_us"] / full_decode_us * 100.0 if full_decode_us > 0 else 0.0
        values.append(
            [
                row["layer_group"],
                row["module"],
                row["kernel"],
                row["stream_no"],
                float(row["time_us"]),
                percent,
            ]
        )
    return values


def write_decode_csv(path: str, values: list[list[Any]]) -> None:
    """Write CSV with repeated group/module cells blanked for readability."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(DECODE_BREAKDOWN_HEADER)
        prev_layer_group: str | None = None
        prev_module: str | None = None
        for row in values:
            layer_group = str(row[0])
            module = str(row[1])
            display_layer_group = "" if layer_group == prev_layer_group else layer_group
            display_module = (
                ""
                if layer_group == prev_layer_group and module == prev_module
                else module
            )
            writer.writerow(
                [
                    display_layer_group,
                    display_module,
                    row[2],
                    row[3],
                    f"{row[4]:.3f}",
                    f"{row[5]:.6f}",
                ]
            )
            prev_layer_group = layer_group
            prev_module = module


def merge_same_value_runs(ws: Any, column: int, start_row: int, end_row: int) -> None:
    run_start = start_row
    prev_value = ws.cell(row=start_row, column=column).value
    for row in range(start_row + 1, end_row + 2):
        value = ws.cell(row=row, column=column).value if row <= end_row else None
        if value != prev_value:
            if prev_value not in (None, "") and row - run_start > 1:
                ws.merge_cells(
                    start_row=run_start,
                    start_column=column,
                    end_row=row - 1,
                    end_column=column,
                )
            run_start = row
            prev_value = value


def write_decode_xlsx(path: str, values: list[list[Any]]) -> None:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "decode_breakdown"
    ws.append(DECODE_BREAKDOWN_HEADER)
    for value_row in values:
        display_row = list(value_row)
        kernel = str(display_row[2])
        if len(kernel) > XLSX_KERNEL_DISPLAY_LIMIT:
            display_row[2] = kernel[: XLSX_KERNEL_DISPLAY_LIMIT - 3] + "..."
        ws.append(display_row)
        if len(kernel) > XLSX_KERNEL_DISPLAY_LIMIT:
            ws.cell(row=ws.max_row, column=3).comment = Comment(kernel, "ATOM")

    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    end_row = ws.max_row
    if end_row >= 2:
        module_run_start = 2
        prev_key = (ws.cell(row=2, column=1).value, ws.cell(row=2, column=2).value)
        for row in range(3, end_row + 2):
            key = (
                ws.cell(row=row, column=1).value if row <= end_row else None,
                ws.cell(row=row, column=2).value if row <= end_row else None,
            )
            if key != prev_key:
                if prev_key[1] not in (None, "") and row - module_run_start > 1:
                    ws.merge_cells(
                        start_row=module_run_start,
                        start_column=2,
                        end_row=row - 1,
                        end_column=2,
                    )
                module_run_start = row
                prev_key = key

        merge_same_value_runs(ws, 1, 2, end_row)

    for row in ws.iter_rows(min_row=2):
        is_total_row = row[2].value == "GROUP TOTAL"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_total_row:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        row[4].number_format = "0.000"
        row[5].number_format = "0.000000"

    widths = {
        1: 24,
        2: 72,
        3: 72,
        4: 10,
        5: 12,
        6: 28,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(path)


def write_decode_output(
    path: str,
    rows: list[dict[str, Any]],
    full_decode_us: float,
    print_head: bool = False,
) -> None:
    stream_map = remap_streams([{"args": {"stream": row["stream"]}} for row in rows])
    breakdown_rows = build_grouped_breakdown_rows(
        rows, stream_map, print_head=print_head
    )
    values = decode_breakdown_values(breakdown_rows, full_decode_us)
    if path.lower().endswith(".xlsx"):
        write_decode_xlsx(path, values)
    else:
        write_decode_csv(path, values)


def main() -> None:
    parser = argparse.ArgumentParser(description="Formal ATOM trace parser")
    parser.add_argument("run_trace")
    parser.add_argument("--capture-trace", default=None)
    parser.add_argument(
        "--output",
        default="decode_breakdown.xlsx",
        help="Output path, .xlsx or .csv (default: decode_breakdown.xlsx).",
    )
    parser.add_argument(
        "--kernel-num",
        type=int,
        default=100,
        help="Number of warmup kernel mappings to print (default: 100).",
    )
    parser.add_argument(
        "--print-head",
        action="store_true",
        help="Include non-layer prologue/epilogue rows in the output.",
    )
    args = parser.parse_args()

    run_events = load_events(args.run_trace)
    capture_trace = args.capture_trace or find_capture_trace(args.run_trace)
    if capture_trace is None:
        raise RuntimeError(
            "Could not auto-discover capture trace; pass --capture-trace."
        )
    capture_events = load_events(capture_trace)

    decode = find_first_decode(run_events)
    batch_size = decode_batch_size(decode)
    graph = find_capture_graph_for_bs(capture_events, batch_size)
    warmup_start, warmup_end = warmup_window_for_graph(capture_events, graph)
    counts = count_events_in_window(capture_events, warmup_start, warmup_end)

    print(f"Run trace: {args.run_trace}")
    print(f"Capture trace: {capture_trace}")
    print("")
    print("First decode:")
    print(f"  name: {decode.get('name')}")
    print(f"  ts: {decode.get('ts'):.3f}")
    print(f"  dur: {decode.get('dur'):.3f}")
    print(f"  batch size: {batch_size}")
    print("")
    print("Matching capture graph:")
    print(f"  name: {graph.get('name')}")
    print(f"  ts: {graph.get('ts'):.3f}")
    print(f"  dur: {graph.get('dur'):.3f}")
    print("")
    print("Decode warmup window:")
    print(f"  start: {warmup_start:.3f}")
    print(f"  end: {warmup_end:.3f}")
    print(f"  dur: {warmup_end - warmup_start:.3f}")
    print(f"  events: {counts}")
    warmup_mapping = build_warmup_mapping(capture_events, warmup_start, warmup_end)
    print(f"  mapping entries: {len(warmup_mapping)}")
    print_first_warmup_mappings(warmup_mapping, limit=args.kernel_num)

    decode_start, decode_end = decode_gpu_window(run_events, decode)
    replay_kernels = replay_kernels_in_window(run_events, decode_start, decode_end)
    matched_rows = match_replay_to_warmup(replay_kernels, warmup_mapping)
    full_decode_us = decode_end - decode_start
    write_decode_output(
        args.output,
        matched_rows,
        full_decode_us,
        print_head=args.print_head,
    )
    unmatched = sum(1 for row in matched_rows if row["cpu_module"] == "<unmatched>")
    print("")
    print("Decode replay mapping:")
    print(f"  replay kernels: {len(replay_kernels)}")
    print(f"  unmatched kernels: {unmatched}")
    print(f"  output written to: {args.output}")


if __name__ == "__main__":
    main()
